#!/usr/bin/env python3
"""
One-time importer: AISEHack registration-form .xlsx -> .mail-ops-cache/teams.json

Run this whenever there's a fresh export of the registration Google Form
responses, so the mail-ops dashboard's Teams view and "who is this sender"
lookup have current data. Requires `pip install openpyxl`.

Usage:
    python3 scripts/import-teams.py "/path/to/Registration form ... (Responses).xlsx"

Output is gitignored (.mail-ops-cache/) — it contains real students' names,
phone numbers, and emails and must never be committed.
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing dependency: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

TRACK_MAP = {
    "Remote sensing (Contributed by Galaxeye)": "sar",
    "Molecular Property prediction (Contributed by IIT Madras)": "polymer",
}


def track_id(raw: str) -> str:
    return TRACK_MAP.get((raw or "").strip(), "unknown")


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/import-teams.py <path-to-xlsx>", file=sys.stderr)
        sys.exit(1)

    src = Path(sys.argv[1]).expanduser()
    if not src.exists():
        print(f"File not found: {src}", file=sys.stderr)
        sys.exit(1)

    repo_root = Path(__file__).resolve().parent.parent
    out_dir = repo_root / ".mail-ops-cache"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "teams.json"

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    teams = []
    by_email = {}

    if "Form Responses 1" in wb.sheetnames:
        ws = wb["Form Responses 1"]
        rows = list(ws.iter_rows(values_only=True))
        for i, r in enumerate(rows[1:]):
            if not any(r):
                continue
            if len(r) < 23:
                r = r + (None,) * (23 - len(r))
            members = []
            for slot in range(5):
                name = r[8 + slot]
                email = r[13 + slot]
                kaggle = r[18 + slot]
                if not (name or email):
                    continue
                members.append(
                    {
                        "name": (name or "").strip() if isinstance(name, str) else name,
                        "email": (email or "").strip().lower() if isinstance(email, str) else email,
                        "kaggleId": (kaggle or "").strip() if isinstance(kaggle, str) else kaggle,
                    }
                )

            team = {
                "id": f"team-{i}",
                "timestamp": r[0].isoformat() if hasattr(r[0], "isoformat") else str(r[0]),
                "track": track_id(r[5]),
                "trackRaw": r[5],
                "soloOrTeam": r[7],
                "willingToTravel": r[6],
                "place": r[4],
                "contactNumber": str(int(r[2])) if isinstance(r[2], float) else (str(r[2]) if r[2] is not None else ""),
                "linkedin": r[1],
                "twitter": r[3],
                "members": members,
            }
            teams.append(team)
            for m in members:
                if m["email"]:
                    by_email[m["email"]] = team["id"]

    submissions = []
    if "Round 1 Submissions" in wb.sheetnames:
        ws = wb["Round 1 Submissions"]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows[1:]:
            if not any(r):
                continue
            if len(r) < 5:
                r = r + (None,) * (5 - len(r))
            name = r[0]
            if isinstance(name, str):
                name = name.strip()
            submissions.append(
                {
                    "teamName": name,
                    "track": r[1],
                    "kaggleNotebook": r[2],
                    "code": r[3],
                    "rank": r[4],
                }
            )

    data = {
        "generatedAt": __import__("datetime").datetime.now().isoformat(),
        "sourceFile": src.name,
        "teams": teams,
        "emailIndex": by_email,
        "submissions": submissions,
    }

    out_path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {len(teams)} teams ({len(by_email)} member emails indexed), {len(submissions)} submissions -> {out_path}")


if __name__ == "__main__":
    main()
